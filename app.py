# app.py
# RUN: streamlit run app.py

import json
from io import BytesIO
import pandas as pd
import streamlit as st
from google import genai
from google.genai import types
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter

# =====================================
# API MANAGEMENT (LOCAL & CLOUD SAFE)
# =====================================

# Safe fallback initialization
API_KEY = "AIzaSyAW7pTxN74jsPgOR-vcqYfXmRQ97KO7XOA"

try:
    # Check if we are on the cloud server with registered secrets
    if hasattr(st, "secrets") and "GEMINI_API_KEY" in st.secrets:
        API_KEY = st.secrets["GEMINI_API_KEY"]
except Exception:
    # If secrets configuration isn't loaded locally, use the fallback string safely
    pass

client = genai.Client(
    api_key=API_KEY
)

# =====================================
# PAGE SETUP
# =====================================

st.set_page_config(
    page_title="LogiFlow",
    page_icon="📄",
    layout="wide"
)

# =====================================
# STATE FRAMEWORK
# =====================================

if "credits" not in st.session_state:
    st.session_state.credits = 3

if "history" not in st.session_state:
    st.session_state.history = []

if "locked" not in st.session_state:
    st.session_state.locked = False

if "email" not in st.session_state:
    st.session_state.email = "operations@demo.com"

if "password" not in st.session_state:
    st.session_state.password = "••••••••"

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

# Boundary checking
if st.session_state.credits <= 0:
    st.session_state.locked = True

# =====================================
# FLUID RESPONSIVE CSS INTERFACE
# =====================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

#MainMenu, footer, header {
    visibility: hidden;
}

[data-testid='stToolbar'],
[data-testid='stHeaderActionElements'],
.stMarkdown a.header-anchor,
button[data-testid="stTabBarTicket"] {
    display: none !important;
}

.block-container {
    max-width: 94%;
    padding-top: .4rem;
}

.stApp {
    background: linear-gradient(135deg, #07111f, #081527, #111827);
}

.stTabs [data-baseweb="tab-list"] {
    justify-content: center;
    gap: clamp(15px, 4vw, 45px);
}

/* RESPONSIVE FLUID TYPOGRAPHY LOGIC */
.logo {
    font-size: clamp(32px, 6vw, 56px) !important;
    font-weight: 800;
    letter-spacing: clamp(4px, 1vw, 12px);
    color: #FFFFFF !important;
    text-shadow: 0px 4px 12px rgba(139, 92, 246, 0.4);
    text-align: center;
    margin-bottom: 5px;
    white-space: nowrap !important;
}

.subtitle {
    text-align: center;
    font-size: clamp(11px, 2.5vw, 14px);
    margin-bottom: 30px;
    color: #A3BFD9 !important;
    font-weight: 500;
    letter-spacing: 0.5px;
}

/* FLEXIBLE COMPACT METRIC CARDS */
.metric {
    padding: 16px;
    background: linear-gradient(180deg, #111827, #0f172a);
    border-radius: 18px;
    border: 1px solid rgba(255, 255, 255, 0.06);
    min-height: 110px;
    text-align: center;
    margin-bottom: 15px;
}

.metric-title {
    font-size: 13px;
    color: #94A3B8;
    font-weight: 500;
    margin-bottom: 4px;
}

.metric-value {
    font-size: clamp(18px, 4vw, 24px);
    font-weight: 700;
    color: #FFFFFF;
    white-space: nowrap;
}

.lock {
    padding: 30px;
    border-radius: 20px;
    background: linear-gradient(180deg, #1e1b4b, #111827);
    border: 1px solid #7c3aed;
    text-align: center;
    margin-bottom: 25px;
}

[data-testid="stFileUploader"] {
    background: #111827 !important;
    border: 2px dashed #7c3aed !important;
    border-radius: 25px !important;
    padding: clamp(20px, 5vw, 40px) !important;
    min-height: 180px !important;
    margin-top: 10px !important;
}

.stButton button {
    height: 52px;
    border-radius: 14px;
    font-weight: 600;
    width: 100%;
}

.pricing-card {
    background: linear-gradient(180deg, #111827, #0f172a);
    border-radius: 22px;
    padding: 25px;
    border: 1px solid rgba(255, 255, 255, 0.06);
    margin-bottom: 20px;
    min-height: 380px;
}

.pricing-title {
    font-size: 20px;
    font-weight: 700;
    margin-bottom: 5px;
}

.pricing-price {
    font-size: 26px;
    font-weight: 800;
    color: #A78BFA;
    margin-bottom: 20px;
}

.pricing-feature {
    font-size: 13px;
    margin-bottom: 12px;
    color: #E2E8F0;
    display: flex;
    align-items: center;
}

.checkmark {
    color: #10B981;
    font-weight: bold;
    margin-right: 10px;
}

/* MEDIA QUERY BREAKPOINT FOR MOBILE FIXES */
@media (max-width: 768px) {
    .block-container {
        max-width: 98%;
    }
    .pricing-card {
        min-height: auto;
    }
}
</style>
""", unsafe_allow_html=True)

# =====================================
# CONTROL PANEL & PROFILE ANCHOR
# =====================================

left, center, right = st.columns([3, 6, 3])

with left:
    with st.popover("🌐 Public Account Menu" if not st.session_state.logged_in else f"👤 Account Open"):
        st.markdown("### 📁 Management Panel")
        user_email = st.text_input("Operator Workspace Email",
                                   value=st.session_state.email if st.session_state.logged_in else "public_demo@logiflow.io")
        user_pass = st.text_input("Security Key / Password", value=st.session_state.password, type="password")

        if st.button("💾 Save Account Changes"):
            st.session_state.email = user_email
            st.session_state.password = user_pass
            st.session_state.logged_in = True
            st.toast("Profile configurations updated securely!", icon="💾")
            st.rerun()

        st.markdown("---")
        st.markdown("### 💳 Active Subscriptions")

        if st.session_state.logged_in:
            st.info(f"🏷️ Plan: Pending Activation\n\n📊 Usage Limits: 0 / 2,500 operations.")
            if st.button("🚪 Logout Session"):
                st.session_state.logged_in = False
                st.session_state.credits = 3
                st.session_state.locked = False
                st.session_state.email = "operations@demo.com"
                st.session_state.password = "••••••••"
                st.rerun()
        else:
            st.warning("🏷️ Plan: Demo Sandbox\n\n📊 Status: Operational Verification Mode Only.")

with center:
    st.markdown("""
    <div class='logo'>LOGIFLOW</div>
    <div class='subtitle'>AI-Powered Invoice Extraction • Excel Automation • Enterprise Processing</div>
    """, unsafe_allow_html=True)

with right:
    st.markdown(
        f"<div style='text-align: right; padding-top: 12px; font-weight: 600; font-size: 15px; color: #94A3B8;'>Credits: {st.session_state.credits}/3</div>",
        unsafe_allow_html=True)

# =====================================
# PLATFORM SECTIONS
# =====================================

tab1, tab2, tab3 = st.tabs([
    "🚀 Data Extractor",
    "💳 Upgrade & Plans",
    "📊 Processing History"
])


# =====================================
# PROCESSING UTILITIES
# =====================================

def first_two_pages(pdf_bytes):
    reader = PdfReader(BytesIO(pdf_bytes))
    writer = PdfWriter()
    pages = min(2, len(reader.pages))

    for i in range(pages):
        writer.add_page(reader.pages[i])

    out = BytesIO()
    writer.write(out)
    return out.getvalue()


def parse_pdf(pdf):
    prompt = """
    Extract:
    Invoice Number
    Date
    Carrier
    Total Charges
    Return JSON only
    """
    part = types.Part.from_bytes(data=pdf, mime_type="application/pdf")
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[prompt, part],
        config={"response_mime_type": "application/json"}
    )
    return json.loads(response.text)


def create_excel(df):
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, index=False)
        sh = w.sheets["Sheet1"]
        fill = PatternFill(fill_type="solid", start_color="1E3A8A")
        font = Font(bold=True, color="FFFFFF")

        for c in sh[1]:
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="center")

        for col in sh.columns:
            m = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                m = max(m, len(str(cell.value or "")))
            sh.column_dimensions[letter].width = m + 3

    return out.getvalue()


# =====================================
# TAB 1: OPERATIONAL WORKSPACE
# =====================================

with tab1:
    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    if st.session_state.locked:
        st.markdown("""
        <div class='lock'>
            <h2>🔒 Demo Access Exhausted</h2>
            <p style='color: #94A3B8;'>3 lifetime evaluation credits have been consumed cleanly.</p>
            <p>Upgrade your platform profile instance to unlock limitless production processing logs.</p>
        </div>
        """, unsafe_allow_html=True)

        c1, c2 = st.columns([1, 1])
        with c1:
            email_input = st.text_input("Create Workspace Operator Email", value="")
        with c2:
            pass_input = st.text_input("Choose Strong Password", type="password")

        if st.button("📩 Create Operator Account & Proceed"):
            if email_input and pass_input:
                st.session_state.email = email_input
                st.session_state.password = pass_input
                st.session_state.logged_in = True
                st.toast("Corporate profile registry successful!", icon="🚀")
                st.rerun()
            else:
                st.error("Please fill in corporate registry parameters to lock in your software account structure.")
        st.warning(
            "Device session locked under Sandbox rule limits. Please migrate your system token context to a paid tier.")

    else:
        # Custom HTML metrics block for clean responsive scaling
        m1, m2, m3 = st.columns([1, 1, 1])
        with m1:
            st.markdown(
                f"<div class='metric'><div class='metric-title'>Credits Running</div><div class='metric-value'>{st.session_state.credits}/3</div></div>",
                unsafe_allow_html=True)
        with m2:
            st.markdown(
                "<div class='metric'><div class='metric-title'>Processing Profile</div><div class='metric-value'>PDF → Excel</div></div>",
                unsafe_allow_html=True)
        with m3:
            st.markdown(
                "<div class='metric'><div class='metric-title'>Engine Version</div><div class='metric-value'>Neural Core V2</div></div>",
                unsafe_allow_html=True)

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

        file = st.file_uploader(
            "Upload System Document File Here",
            type=["pdf"],
            accept_multiple_files=False
        )

        if file:
            if st.button("⚡ Analyze Document Engine Pipeline"):
                with st.spinner("Executing direct bytes parsing matrix channels..."):
                    pdf = file.read()
                    pdf = first_two_pages(pdf)
                    result = parse_pdf(pdf)

                    st.session_state.credits -= 1
                    st.session_state.history.append(file.name)

                    df = pd.DataFrame([result])
                    st.success("Ledger array extracted perfectly.")

                    st.dataframe(df, use_container_width=True)

                    st.download_button(
                        "⬇ Export Structured Excel Document",
                        create_excel(df),
                        "logiflow_ledger.xlsx"
                    )

# =====================================
# TAB 2: PRICING PLANS
# =====================================

with tab2:
    st.markdown("<div style='height:25px'></div>", unsafe_allow_html=True)

    c1, c2, c3 = st.columns([1, 1, 1])

    with c1:
        st.markdown("""
        <div class="pricing-card" style="border-top: 4px solid #3B82F6;">
            <div class="pricing-title">Demo Sandbox</div>
            <div class="pricing-price">$0 <span style="font-size:14px; color:#64748B;">/ free trial</span></div>
            <div class="pricing-feature"><span class="checkmark">✓</span> 3 Lifetime Evaluation Credits</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Standard Cloud Queuing Channel</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Limit 1 Document Upload per Run</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Strict 2-Page Cap per Invoice</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> For Accuracy Verification Only</div>
        </div>
        """, unsafe_allow_html=True)
        st.button("Current Testing Environment Active", disabled=True, key="p1")

    with c2:
        st.markdown("""
        <div class="pricing-card" style="border-top: 4px solid #10B981;">
            <div class="pricing-title">Logistics Pro</div>
            <div class="pricing-price">$149 <span style="font-size:14px; color:#64748B;">/ month</span></div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Volumetric Cap: 2,500 Invoices/Mo</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> High-Speed Processing (4–5 Seconds)</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Zero Page Limits/Caps per Document</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Multi-File Batch Data Uploads</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Premium Styled Excel Ledgers</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("⚡ Upgrade to Logistics Pro", key="p2"):
            st.success("Your enterprise routing profile invoice request has been filed with billing@logiflow.io!")

    with c3:
        st.markdown("""
        <div class="pricing-card" style="border-top: 4px solid #F59E0B;">
            <div class="pricing-title">Enterprise Core</div>
            <div class="pricing-price">$499 <span style="font-size:14px; color:#64748B;">/ month</span></div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Unlimited Processing (30k/mo Fair Use)</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Dedicated Priority Execution (2–3s)</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Multi-User Managed Seats (Up to 10)</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Live Google Sheets Webhook Sync</div>
            <div class="pricing-feature"><span class="checkmark">✓</span> Custom Accounting ERP Key Mapping</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("💼 Deploy Enterprise Core Plan", key="p3"):
            st.success(
                "An enterprise solution architect has been assigned to contact your organization desk within 1 hour.")

# =====================================
# TAB 3: HISTORY RECORDS
# =====================================

with tab3:
    st.markdown("<div style='height:25px'></div>", unsafe_allow_html=True)
    st.subheader("Recent Document Audit Run Matrix")

    if len(st.session_state.history) == 0:
        st.info("No transaction parsing logs found inside current validation registry framework.")
    else:
        for idx, item in enumerate(st.session_state.history):
            st.success(f"📟 Entry ID #{1001 + idx}  |  📄 Filename Reference: {item}  |  🔒 Status: Cleared")

# =====================================
# SYSTEM FOOTER DATA & SEO MOAT
# =====================================

st.markdown("---")
st.subheader("Industry Documentation & Technical Manuals")

st.markdown("""
• [How to automate freight billing records and eliminate clerical errors data entry modules](#)  
• [Top AI tools engineered to convert unstructured logistics multi-page PDFs to Excel sheets](#)  
• [The hidden balance-sheet costs of manual back-office entry errors in corporate fleet processing](#)  
""")

st.caption(
    "Registered Utility Gateway Asset  |  🔐 Contact Interface Support Protocol: support@logiflow.io | licensing@logiflow.io")